from openpyxl.styles import Font, PatternFill

def format_sheet(ws):
    # Format headers: bold, font size 14, remove underscores
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)
        cell.value = str(cell.value).replace('_', ' ')

    # Find column indices
    headers = [cell.value for cell in ws[1]]
    delta_idx = headers.index('price delta 7d(%)') + 1
    printing_idx = headers.index('printing') + 1

    # Apply formatting to each row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # price_delta_7d(%)
        delta_cell = row[delta_idx - 1]
        try:
            delta_val = float(delta_cell.value)
            if delta_val < 0:
                delta_cell.font = Font(color="FF0000")  # Red
            elif delta_val > 0:
                delta_cell.font = Font(color="008000")  # Green
            else:
                delta_cell.font = Font(color="000000")  # Black
        except (TypeError, ValueError):
            pass

        # printing
        printing_cell = row[printing_idx - 1]
        printing_val = str(printing_cell.value).lower()
        if printing_val == "normal":
            printing_cell.font = Font(color="000000")  # Black
        elif printing_val == "rainbow foil":
            # Rainbow: set fill as a gradient (not supported), so use yellow as example
            printing_cell.font = Font(color="FF9900")  # Orange-ish
        elif printing_val == "cold foil":
            printing_cell.font = Font(color="0000FF")  # Blue